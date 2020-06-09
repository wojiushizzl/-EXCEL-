from random import randint, choice

import openpyxl

from openpyxl import Workbook

from openpyxl.styles import Font, colors, Border, Side

from openpyxl.drawing.image import Image
# 创建工作簿

wb = Workbook()

# 获取活动工作表

ws = wb.active

# 修改工作表标题

ws.title = '测试'

# 字段标题

ws.append([' ', '语文', '数学', '英语', '总分'])

# 姓名信息

names = '一二三四五六七八九'

lastNames = '赵钱孙李周吴郑王'

for i, c in enumerate(names):

    cell = 'A' +str( i +2)

    ws[cell] = choice(lastNames) + c

# 随机成绩数据

for row in range(2, len(names ) +2):

    # 随机生成没人每门课的成绩

    row = str(row)

    for col in 'BCD':

        ws[col +row] = randint(30, 100)

    # 使用公式计算每个人的总分

    ws['E' +row] = '=sum(B' + row + ':D' + row + ')'


print("B1",ws['B1'])

# 插入图片

ws.add_image(Image('EVA_0.jpg'), 'F1')

# 合并单元格

lastRow = str(len(names ) +2)

ws.merge_cells('B' +lastRow +':I' +lastRow)

ws['A' +lastRow] = '说明:'

ws['B' +lastRow] = '这只是个测试。'

# 通用边框信息

left, right, top, bottom = [Side(style='medium' ,color='000000') ] *4

border = Border(left=left, right=right, top=top, bottom=bottom)

# 设置单元格边框和颜色

# 表头和最后一行的说明使用默认的黑色

for row in range(2, len(names ) +2):

    # 奇偶行字体颜色交替

    if row %2 == 0:

        color = colors.RED

    else:

        color = '00CCFF'

    # 设置边框和颜色

    for col in 'ABCDE':

        ws[col +str(row)].border = border

        ws[col +str(row)].font = Font(color=color)

for row in ws.rows:
    for cell in row:
        cell.fill = openpyxl.styles.fills.GradientFill(stop=['0000FF', '0000FF'])

wb.save('测试.xlsx')